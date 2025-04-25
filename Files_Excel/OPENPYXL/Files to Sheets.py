import os, sys, shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
import _functions as cfx

choice = int(cfx.inputbox(title='Method', prompt="Choose the method : \n   1. Combined Same Sheet's Data\n   2. Dont Combine Same Sheet's Data [FileName_SheetName] "))

if choice == 1:
    date_style = NamedStyle(name="date_style", number_format='MM/DD/YYYY')
    folder_path = cfx.ifolder()
    new_wb = Workbook()
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    for file_name in excel_files:
        file_path = os.path.join(folder_path, file_name)
        current_wb = load_workbook(file_path)
        for sht in current_wb.sheetnames:
            sheet = current_wb[sht]
            if sht in new_wb.sheetnames:
                new_sheet = new_wb[sht]
                if new_sheet.max_row == 1:
                    start_row = new_sheet.max_row
                    data_start_row = 2
                else:
                    start_row = new_sheet.max_row + 1
                    data_start_row = 2
            else:
                new_sheet = new_wb.create_sheet(title=sht)
                start_row = 1
                data_start_row = 1

            for row in sheet.iter_rows(min_row=data_start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for col_num, cell in enumerate(row, start=1):
                    if isinstance(cell.value, datetime):
                        new_sheet.cell(row=start_row, column=col_num, value=cell.value).style = date_style
                    else:
                        new_sheet.cell(row=start_row, column=col_num, value=cell.value)
                start_row += 1

    if "Sheet" in new_wb.sheetnames:
        del new_wb["Sheet"]
            
    ofolder_path = os.path.join(folder_path, 'Files_Sheets')

    if os.path.exists(ofolder_path):
        shutil.rmtree(ofolder_path)
    os.makedirs(ofolder_path)

    new_wb.save(os.path.join(ofolder_path, 'Files_Sheets.xlsx'))
    new_wb.close()
    os.startfile(folder_path)

elif choice == 2:
    date_style = NamedStyle(name="date_style", number_format='MM/DD/YYYY')
    folder_path = cfx.ifolder()
    new_wb = Workbook()
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    is_first_sheet = True
    for file_name in excel_files:
        file_path = os.path.join(folder_path, file_name)
        current_wb = load_workbook(file_path)
        for sht in current_wb.sheetnames:
            current_row = 1
            sheet = current_wb[sht]
            sh_nm = os.path.splitext(file_name)[0] + "_" + sht
            new_sheet = new_wb.create_sheet(title=sh_nm)
            for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column), start=1):
                data = [cell.value for cell in row]
                for col_num, value in enumerate(data, start=1):
                    if isinstance(value, datetime):
                        new_sheet.cell(row=current_row, column=col_num, value=value).style = date_style
                    else:
                        new_sheet.cell(row=current_row, column=col_num, value=value)
                current_row += 1
            is_first_sheet = False

    if "Sheet" in new_wb.sheetnames:
        del new_wb["Sheet"]

    ofolder_path = os.path.join(folder_path, 'Files_Sheets')

    if os.path.exists(ofolder_path):
        shutil.rmtree(ofolder_path)
    os.makedirs(ofolder_path)

    new_wb.save(os.path.join(ofolder_path, 'Files_Sheets.xlsx'))
    new_wb.close()
    os.startfile(folder_path)

else:
    sys.exit(0)