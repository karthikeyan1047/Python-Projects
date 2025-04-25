import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
import _functions as cfx

folder_path = cfx.ifolder()
workbook_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xls')]

combined_workbook = Workbook()
combined_sheet = combined_workbook.active
combined_sheet.title = 'Combined'
current_row = 1
is_first_sheet = True
date_style = NamedStyle(name="date_style", number_format='MM/DD/YYYY')
for file in workbook_files:
    file_path = os.path.join(folder_path, file)
    try:
        wb = load_workbook(file_path)
    except:
        continue
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]     
        for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column), start=1):
            if row_num == 1 and not is_first_sheet:
                continue
            data = [cell.value for cell in row]
            for col_num, value in enumerate(data, start=1):
                if isinstance(value, datetime):
                    combined_sheet.cell(row=current_row, column=col_num, value=value).style = date_style
                else:    
                    combined_sheet.cell(row=current_row, column=col_num, value=value)            
            current_row += 1
        is_first_sheet = False
    wb.close()
combined_workbook.save(os.path.join(folder_path, 'Files_Sheet.xlsx'))
