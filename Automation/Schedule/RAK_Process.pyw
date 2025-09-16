from datetime import datetime
from easygui import *
import os, pandas, time, shutil
import tkinter as tk
import xlwings as xw
import openpyxl
from openpyxl.styles import Side, Border, Font, PatternFill
import pandas as pd
from tkinter import filedialog, messagebox, simpledialog

dwl_folder = r"C:\Users\karthikeyans\Downloads"
excel_file = r"C:\Users\karthikeyans\Documents\BLUMIN\Automations\WebAutomation_ProcessMed.xlsx"
base_folder = r"C:\Users\karthikeyans\Desktop\RAK Downloaded Files"
his_folder = f"{base_folder}\\HIS"
fpa_folder = f"{base_folder}\\FPA"
lw_folder = f"{base_folder}\\Linewise"
empty_data = pandas.DataFrame()

yr_curr = datetime.today().year

def close_book(file_path):
    wb = xw.Book(file_path)
    wb.save()
    wb.close()
    xw.App().quit()

def clearing(sheet):
    try:
        for row in sheet[sheet.min_row:sheet.max_row]:
            for cell in row:
                cell.value = None
                cell.fill = PatternFill()
                cell.border = Border()
                cell.font = Font()
    except:
        cell = sheet['A1']
        cell.value = None
        cell.fill = PatternFill()
        cell.border = Border()
        cell.font = Font()

def autofit_columns(sheet):
    last_row = sheet.max_row
    last_col = sheet.max_column
    for col in range(1, last_col + 1):
        max_length = 0
        column = openpyxl.utils.get_column_letter(col)
        for row in range(1, last_row + 1):
            cell = sheet[f'{column}{row}']
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    if "_" in str(cell.value):
                        max_length = len(str(cell.value)) + 2
                    else:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 3
        sheet.column_dimensions[column].width = adjusted_width

def show_info(title, prompt):
    root = tk.Tk()
    root.attributes('-topmost', True)

    window_width = 300
    window_height = 150
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    root.withdraw()
    root.withdraw()
    messagebox.showinfo(title, prompt)
    root.quit()

def inputbox(title, prompt):
    ttl = title
    pmt = prompt
    root = tk.Tk()
    root.attributes('-topmost', True)

    window_width = 300
    window_height = 150
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    root.withdraw()
    user_input = simpledialog.askstring(ttl, pmt)
    return user_input

def open_base_folder():
    os.startfile(base_folder)

def generate_month_ranges(year, step):
    current_year = datetime.today().year
    current_month = datetime.today().month
    ranges = []

    for i in range(1, 13, step):
        start = i
        end = min(i + step - 1, 12)

        if year < current_year:
            ranges.append((start, end))
        elif year == current_year:
            if start > current_month:
                break
            ranges.append((start, min(end, current_month)))
    
    return ranges

def renaming_files():
    close_book(excel_file)

    choice = int(inputbox(title="RAK/SOA", prompt="1. RAK\n2. SOA\n3. Reconciliation"))
    mapping = {
        1 : "RAK_Rename",
        2 : "SOA_Rename",
        3 : "Recon_Rename"
    }
    sheet_name = mapping.get(choice, "Invalid choice")
    df = pd.read_excel(excel_file, sheet_name=sheet_name)

    rename_dict = dict(df)

    old_names = rename_dict['Old Name'].to_list()
    new_names = rename_dict['New Name'].to_list()
    status = rename_dict['Status'].to_list()

    for old_name, new_name, status in zip(old_names, new_names, status):
        if status == "successfully":
            old_file_path = os.path.join(dwl_folder, f"{old_name}.xlsx")
            new_file_path = os.path.join(dwl_folder, f"{new_name}.xlsx")
            if os.path.exists(old_file_path):
                os.rename(old_file_path, new_file_path)

    os.startfile(dwl_folder)
    
def delete_old_files(folders):
    for folder in folders:
        for dirpath, dirnames, filenames in os.walk(folder):
            for filename in filenames:
                file_path = os.path.join(dirpath, filename)
                os.remove(file_path)

    open_folder = f"{fpa_folder}\\2023\\DHPO"
    os.startfile(open_folder)


def downloads_to_rak():
    f_count = 0
    for file_name in os.listdir(dwl_folder):
        src_file_path = os.path.join(dwl_folder, file_name)
        if ".xlsx" in file_name:
            parts = file_name.replace('.xlsx', '').split('_')

            template, year, instance = parts[0], parts[3], parts[1]

            if template in ['Sub', 'Resub']:
                folder_name = f"{template}_RA"
                if template == 'Sub':
                    dest_paths = [os.path.join(base_folder, 'FPA', year, instance, file_name),
                            os.path.join(base_folder, 'Linewise', year, instance, folder_name, file_name)
                            ]
                else:
                    dest_paths = [os.path.join(base_folder, 'Linewise', year, instance, folder_name, file_name)]
            else:
                dest_paths = [os.path.join(base_folder, template, year, instance, file_name)]

            for dest_path in dest_paths:
                # print(src_file_path, ' : ', dest_path)
                shutil.copy2(src_file_path, dest_path)
                f_count += 1

    show_info("Total Files", f"Total Files copied : {f_count}")
    open_folder = f"{fpa_folder}\\2023\\DHPO"
    os.startfile(open_folder)

def rak_to_source():
    bat_file = r"C:\Users\karthikeyans\Documents\BLUMIN\Automations\RAK File Transfer\1.RAK.bat"
    os.startfile(bat_file)

def delete_dwl_files():
    close_book(excel_file)

    choice = int(inputbox(title="RAK/SOA", prompt="1. RAK\n2. SOA\n3. Reconciliation"))
    mapping = {
        1 : "RAK_Rename",
        2 : "SOA_Rename",
        3 : "Recon_Rename"
    }
    sheet_name = mapping.get(choice, "Invalid choice")
    df = pd.read_excel(excel_file, sheet_name=sheet_name)

    name_dict = dict(zip(df['Old Name'], df['New Name']))
    for file_name in os.listdir(dwl_folder):
        for old_name, new_name in name_dict.items():
            if file_name == f"{new_name}.xlsx":
                del_file_path = os.path.join(dwl_folder, file_name)
                os.remove(del_file_path)

    os.startfile(dwl_folder)

rak_folders = [his_folder, fpa_folder, lw_folder]

def rak_process():
    root = tk.Tk()
    root.title("File Operations")
    root.attributes('-topmost', True)
    window_width = 300
    window_height = 300

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = int((screen_width / 2) - (window_width / 2))
    y = int((screen_height / 2) - (window_height / 2))
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")

    tk.Label(root, text="Choose :", font=('Copperplate Gothic Bold', 12, 'bold'), fg='black').pack(pady=5)

    button_frame = tk.Frame(root)
    button_frame.pack()

    tk.Button(button_frame, text="Open Base Folder", width=30, font= ('Bahnschrift SemiLight SemiConde', 12, 'bold'), fg='black', bg='#F08080', command=lambda: [open_base_folder(), root.destroy()]).pack(pady=5)
    tk.Button(button_frame, text="Rename Files", width=30, font= ('Bahnschrift SemiLight SemiConde', 12, 'bold'), fg='black', bg='#F08080', command=lambda: [renaming_files(), root.destroy()]).pack(pady=5)
    tk.Button(button_frame, text="Delete - Old RAK Files", width=30, font= ('Bahnschrift SemiLight SemiConde', 12, 'bold'), fg='black', bg='#F08080', command=lambda: [delete_old_files(rak_folders), root.destroy()]).pack(pady=5)
    tk.Button(button_frame, text="Copy - Downloads to RAK", width=30, font= ('Bahnschrift SemiLight SemiConde', 12, 'bold'), fg='black', bg='#F08080', command=lambda: [downloads_to_rak(), root.destroy()]).pack(pady=5)
    tk.Button(button_frame, text="Move - RAK to Source", width=30, font= ('Bahnschrift SemiLight SemiConde', 12, 'bold'), fg='black', bg='#F08080', command=lambda: [rak_to_source(), root.destroy()]).pack(pady=5)
    tk.Button(button_frame, text="Delete Downloaded Files", width=30, font= ('Bahnschrift SemiLight SemiConde', 12, 'bold'), fg='black', bg='#F08080', command=lambda: [delete_dwl_files(), root.destroy()]).pack(pady=5)

    root.mainloop()

rak_process()