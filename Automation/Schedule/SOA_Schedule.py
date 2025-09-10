from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
import xlwings as xw
import pandas as pd
from datetime import datetime, timedelta
import time
import _functions as cfx
import pyautogui as gui

options = Options()
prefs = {
    "download.prompt_for_download": True,
    "download.default_directory": "",
    "profile.default_content_settings.popups": 1
}
options.add_argument("--start-maximized")
options.add_argument("--disable-popup-blocking")
options.add_argument('--ignore-ssl-errors')
options.add_argument('--ignore-certificate-errors')
options.add_experimental_option("detach", True)
options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(options=options)
driver.implicitly_wait(60)

zoom = '90%'

def wait_for_element(by, value, timeout=20):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))

def wait_for_clickable_element(by, value, timeout=20):
    return WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))

def login(username, password):
    driver.get(url)
    wait_for_element(By.ID, "email").clear()
    driver.find_element(By.ID, "email").send_keys(username)
    driver.find_element(By.ID, "password").clear()
    driver.find_element(By.ID, "password").send_keys(password)
    time.sleep(1)
    wait_for_clickable_element(By.XPATH, "//button[text()='Sign In']").click()
    time.sleep(2)
    driver.execute_script(f"document.body.style.zoom='{zoom}'")
    time.sleep(1)
    wait_for_clickable_element(By.XPATH, "//i[contains(@class, 'pi-bars')]").click()
    time.sleep(1)
    wait_for_clickable_element(By.XPATH, "//i[contains(@class, 'pi-sun')]").click()
    time.sleep(1)

def dropdown_select(dropdown, option):
    dd = wait_for_clickable_element(By.ID, dropdown)
    dd.click()
    opt = wait_for_clickable_element(
        By.XPATH, f"//div[contains(@class, 'p-overlay')]//*[normalize-space()='{option}']"
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", opt)
    opt.click()
    time.sleep(1)

def login_and_select_date_field(username, password, date_field):
    login(username, password)
    dropdown_select('pn_id_17', date_field)

def generate_month_ranges(year, step):
    current_year = datetime.today().year
    current_month = datetime.today().month
    ranges = []

    for i in range(1, 13, step):
        start = i
        end = min(i + step - 1, 12)

        if year < current_year:
            ranges.append((start, end))
        elif year == current_year:
            if start > current_month:
                break
            ranges.append((start, min(end, current_month)))
    
    return ranges

def get_date_range(year, start_month, end_month):
    date_from = datetime(year, start_month, 1)
    if end_month == 12:
        date_to = datetime(year, 12, 31)
    else:
        date_to = datetime(year, end_month + 1, 1) - timedelta(days=1)
    return date_from, date_to

def date_picker(datefrm, dateto):
    def pick_date(date):
        wait = WebDriverWait(driver, 5)
        day = date.day
        month_name, month = date.strftime('%b'), date.month - 1
        year = date.year
        date_label = f"{year}-{month}-{day}"

        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "p-datepicker-select-year"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, f"//span[text()=' {year} ']"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, f"//span[text()=' {month_name} ']"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, f"//span[@data-date='{date_label}']"))).click()

    date_input = driver.find_element(By.CSS_SELECTOR, "p-datepicker input")
    date_input.click()

    pick_date(datefrm)
    pick_date(dateto)

row = 2
def schedule(s_file):
    global row
    schedule_button = wait_for_clickable_element(By.XPATH, "//button[.//span[text()='Schedule']]")
    schedule_button.click()

    WebDriverWait(driver, 300).until(EC.alert_is_present())
    time.sleep(1)

    alert = Alert(driver)
    alert_text0 = alert.text
    alert_text = alert_text0.encode('ascii', errors='replace').decode('ascii')

    if "Please contact administrator" in alert_text:
        file_name, status = 'Error', 'Split and Download'
    elif "No results" in alert_text:
        file_name, status = 'No File', 'No Data'
    else:
        parts = alert_text.replace(",", "").split(" ")
        file_name, status = parts[0], parts[3]

    soa_sh.cell(row=row, column=1).value = file_name
    soa_sh.cell(row=row, column=2).value = s_file
    soa_sh.cell(row=row, column=3).value = status
    row = row+1

    Alert(driver).accept()
    time.sleep(1.5)
    return status

total_files = 0
def process_all_for_user(email, password, centers):
    global driver, total_files, row

    login_and_select_date_field(email, password, date_field)
    dropdown_select('pn_id_7', template)
    for center in centers:
        dropdown_select('pn_id_11', center)
        for year in range(2023, yr_curr + 1):
            month_ranges = generate_month_ranges(year, 4)
            for sm, em in month_ranges:
                date_from, date_to = get_date_range(year, sm, em)
                date_picker(date_from, date_to)
                time.sleep(1)

                mns = '' if len(month_ranges) == 1 else f"_{sm}" if sm == em else f"_{sm}-{em}"
                s_file = f"{center}_{year}{mns}"
                sts = schedule(s_file)
                if sts == 'successfully':
                    total_files += 1
      

workbook_path = r"C:\Users\karthikeyans\Documents\BLUMIN\Automations\WebAutomation_ProcessMed.xlsx"
workbook = load_workbook(workbook_path)
soa_sh_name = 'SOA_Rename'
soa_sh = workbook[soa_sh_name]
cfx.close_book(workbook_path)
cfx.clearing(soa_sh)

url = "https://rcmbi.processmed.ae/login"
username = "hbaybi-dhpo@processmed.ae"
password = "7ofsZit4L2hztw2HKbq^"
date_field = "Encounter.Start"
template = "Healthbay_SOA"
centers = ['HBD', 'HBPV', 'HBPM', 'HBPW']
yr_curr = datetime.today().year
mn_curr = datetime.today().month

soa_sh.cell(row=1, column=1).value = 'Old Name'
soa_sh.cell(row=1, column=2).value = 'New Name'
soa_sh.cell(row=1, column=3).value = 'Status'

process_all_for_user(username, password, centers)
time.sleep(2)

driver.quit()

cfx.show_info("Total Files", f"Total Files generated : {total_files}")

cfx.autofit_columns(soa_sh)
workbook.save(workbook_path)

time.sleep(3)

pg = int((((yr_curr - 2023) + 1) * 3)/10) + 2
df = pd.read_excel(workbook_path, sheet_name='SOA_Rename')
rename_dict = dict(df)
old_names = rename_dict['Old Name'].to_list()

driver = webdriver.Chrome(options=options)
driver.implicitly_wait(60)

login(username, password)

wait_for_clickable_element(By.XPATH, "//button[.//span[text()='Reports']]").click()
for center in centers:
    dropdown_select('clientName', center)
    wait_for_clickable_element(By.XPATH, "//button[.//span[text()='Refresh']]").click()

    wait_for_element(By.XPATH, f"//table//tr/td[2][text()=' {center} ']", 30)

    try:
        wait_for_clickable_element(By.XPATH, "//button[@aria-label='First Page']", 5).click()
    except:
        pass

    for i in range(1, pg+1):
        WebDriverWait(driver, 20).until(
            lambda d: len(d.find_elements(By.XPATH, "//*[@id='pn_id_1-table']/tbody/tr")) >= 10
        )
        rows = driver.find_elements(By.XPATH, "//*[@id='pn_id_1-table']/tbody/tr")
        for r in range(1, len(rows)+1):
            sch_filename = wait_for_element(By.XPATH, f"//*[@id='pn_id_1-table']/tbody/tr[{r}]/td[5]", 5).text
            if sch_filename in old_names:
                wait_for_clickable_element(By.XPATH, f"//*[@id='pn_id_1-table']/tbody/tr[{r}]/td[9]/p-button/button").click()
                time.sleep(2)
                gui.hotkey('enter')
        wait_for_clickable_element(By.XPATH, "//button[@aria-label='Next Page']").click()
driver.quit()