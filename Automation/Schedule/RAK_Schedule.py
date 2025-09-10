from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
import pandas as pd
import xlwings as xw
from datetime import datetime, timedelta
import time, math
import _functions as cfx
import pyautogui as gui


options = Options()
prefs = {
    "download.prompt_for_download": True,
    "download.default_directory": "",
    "profile.default_content_settings.popups": 1
}
options.add_argument("--start-maximized")
options.add_argument("--disable-popup-blocking")
options.add_argument('--ignore-ssl-errors')
options.add_argument('--ignore-certificate-errors')
options.add_experimental_option("detach", True)     # it will not close the browser automatically
options.add_experimental_option("prefs", prefs)

zoom = '90%'

def wait_for_element(by, value, timeout=20):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))

def wait_for_clickable_element(by, value, timeout=20):
    return WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))

def login(username, password):
    driver.get(url)
    wait_for_element(By.ID, "email").clear()
    driver.find_element(By.ID, "email").send_keys(username)
    driver.find_element(By.ID, "password").clear()
    driver.find_element(By.ID, "password").send_keys(password)
    time.sleep(1)
    wait_for_clickable_element(By.XPATH, "//button[text()='Sign In']").click()
    time.sleep(2)
    driver.execute_script(f"document.body.style.zoom='{zoom}'")
    time.sleep(1)
    wait_for_clickable_element(By.XPATH, "//i[contains(@class, 'pi-bars')]").click()
    time.sleep(1)
    wait_for_clickable_element(By.XPATH, "//i[contains(@class, 'pi-sun')]").click()
    time.sleep(1)

def dropdown_select(dropdown, option):
    dd = wait_for_clickable_element(By.ID, dropdown)
    dd.click()
    opt = wait_for_clickable_element(
        By.XPATH, f"//div[contains(@class, 'p-overlay')]//*[normalize-space()='{option}']"
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", opt)
    opt.click()
    time.sleep(1)

def login_and_select_date_field(username, password, date_field):
    login(username, password)
    dropdown_select('pn_id_17', date_field)

def generate_month_ranges(year, step):
    current_year = datetime.today().year
    current_month = datetime.today().month
    ranges = []

    for i in range(1, 13, step):
        start = i
        end = min(i + step - 1, 12)

        if year < current_year:
            ranges.append((start, end))
        elif year == current_year:
            if start > current_month:
                break
            ranges.append((start, min(end, current_month)))
    
    return ranges

def get_date_range(year, start_month, end_month):
    date_from = datetime(year, start_month, 1)
    if end_month == 12:
        date_to = datetime(year, 12, 31)
    else:
        date_to = datetime(year, end_month + 1, 1) - timedelta(days=1)
    return date_from, date_to

def date_picker(datefrm, dateto):
    def pick_date(date):
        wait = WebDriverWait(driver, 5)
        day = date.day
        month_name, month = date.strftime('%b'), date.month - 1
        year = date.year
        date_label = f"{year}-{month}-{day}"

        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "p-datepicker-select-year"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, f"//span[text()=' {year} ']"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, f"//span[text()=' {month_name} ']"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, f"//span[@data-date='{date_label}']"))).click()

    date_input = driver.find_element(By.CSS_SELECTOR, "p-datepicker input")
    date_input.click()

    pick_date(datefrm)
    pick_date(dateto)

row = 2
def schedule(s_file):
    global row
    schedule_button = wait_for_clickable_element(By.XPATH, "//button[.//span[text()='Schedule']]")
    schedule_button.click()

    WebDriverWait(driver, 300).until(EC.alert_is_present())
    time.sleep(1.5)

    alert = Alert(driver)
    alert_text0 = alert.text
    alert_text = alert_text0.encode('ascii', errors='replace').decode('ascii')
    
    if "Please contact administrator" in alert_text:
        file_name, status = 'Error', 'Split and Download'
    elif "No results" in alert_text:
        file_name, status = 'No File', 'No Data'
    else:
        parts = alert_text.replace(",", "").split(" ")
        file_name, status = parts[0], parts[3]

    rak_renames_sh.cell(row=row, column=1).value = file_name
    rak_renames_sh.cell(row=row, column=2).value = s_file
    rak_renames_sh.cell(row=row, column=3).value = status
    row = row+1

    Alert(driver).accept()
    time.sleep(1.5)

    return alert_text, status

file_dict = {
    "Unsettled_Sant" : "HIS",
    "Remittance" : "Sub",
    "Resub_Remittance" : "Resub",
    "RAK" : "H",
    "RAKP" : "P"
}
total_files = 0
def process_all_for_user(email, password, files, instance):
    global driver, total_files
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(60)
    login_and_select_date_field(email, password, date_field)

    for template, yr_start, center, step in files:
        dropdown_select('pn_id_7', template)
        dropdown_select('pn_id_11', center)

        for year in range(yr_start, yr_curr + 1):
            month_ranges = generate_month_ranges(year, step)
            for sm, em in month_ranges:
                date_from, date_to = get_date_range(year, sm, em)
                date_picker(date_from, date_to)
                time.sleep(1)

                mns = '' if len(month_ranges) == 1 else f"_{sm}" if sm == em else f"_{sm}-{em}"
                s_file = f"{file_dict[template]}_{instance}_{file_dict[center]}_{year}{mns}"
                alrt_text, sts = schedule(s_file)

                if "Please contact administrator" in alrt_text:
                    return
                
                if sts == 'successfully':
                    total_files += 1
                
    driver.quit()

workbook_path = r"C:\Users\karthikeyans\Documents\BLUMIN\Automations\WebAutomation_ProcessMed.xlsx"
workbook = load_workbook(workbook_path)
rak_renames_sh_name = "RAK_Rename"
rak_renames_sh = workbook[rak_renames_sh_name]
cfx.close_book(workbook_path)
cfx.clearing(rak_renames_sh)

url = "https://rcmbi.processmed.ae/login"
date_field = "Encounter.End"
yr_curr = datetime.today().year
mn_curr = datetime.today().month

dhpo_email, dhpo_password = 'rakbi-dhpo@processmed.ae', "xL=0UX+:sb=@x#n1X).y"
rpo_email, rpo_password = 'rakbi-rypd@processmed.ae', "kQm?f_8U1uV7Bf>u.By}"
his_year, sub_resub_year = 2022, 2023
dhpo_files = [
    ('Unsettled_Sant', his_year, 'RAK', 6),
    ('Unsettled_Sant', his_year, 'RAKP', 12),
    ('Remittance', sub_resub_year, 'RAK', 4),
    ('Remittance', sub_resub_year, 'RAKP', 12),
    ('Resub_Remittance', sub_resub_year, 'RAK', 6),
    ('Resub_Remittance', sub_resub_year, 'RAKP', 12)
]

rpo_files = [
    ('Unsettled_Sant', his_year, 'RAK', 6),
    ('Unsettled_Sant', his_year, 'RAKP', 12),
    ('Remittance', sub_resub_year, 'RAK', 2),
    ('Remittance', sub_resub_year, 'RAKP', 6),
    ('Resub_Remittance', sub_resub_year, 'RAK', 3),
    ('Resub_Remittance', sub_resub_year, 'RAKP', 12)                
]

rak_renames_sh.cell(row=1, column=1).value = 'Old Name'
rak_renames_sh.cell(row=1, column=2).value = 'New Name'
rak_renames_sh.cell(row=1, column=3).value = 'Status'

process_all_for_user(dhpo_email, dhpo_password, dhpo_files, 'DHPO')
time.sleep(2)
process_all_for_user(rpo_email, rpo_password, rpo_files, 'RPO')
time.sleep(2)

cfx.show_info("Total Files", f"Total Files generated : {total_files}")

cfx.autofit_columns(rak_renames_sh)
workbook.save(workbook_path)

time.sleep(3)

pg = math.ceil(((((yr_curr - 2023) + 1) * 12) + 2) / 10) + 2
login_details = [(dhpo_email, dhpo_password), (rpo_email, rpo_password)]
centers_dwl = ['RAK', 'RAKP']
df = pd.read_excel(workbook_path, sheet_name='RAK_Rename')
rename_dict = dict(df)
old_names = rename_dict['Old Name'].to_list()
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 20)
for idx, (username, password) in enumerate(login_details):
    login(username, password)
    wait_for_clickable_element(By.XPATH, "//button[.//span[text()='Reports']]").click()

    for center in centers_dwl:
        dropdown_select('clientName', center)
        wait_for_clickable_element(By.XPATH, "//button[.//span[text()='Refresh']]").click()

        wait_for_element(By.XPATH, f"//table//tr/td[2][text()=' {center} ']", 30)

        try:
            wait_for_clickable_element(By.XPATH, "//button[@aria-label='First Page']", 5).click()
        except:
            pass

        for i in range(1, pg+1):
            wait.until(
                lambda d: len(d.find_elements(By.XPATH, "//*[@id='pn_id_1-table']/tbody/tr")) >= 10
            )
            rows = driver.find_elements(By.XPATH, "//*[@id='pn_id_1-table']/tbody/tr")
            for r in range(1, len(rows)+1):
                sch_filename = wait_for_element(By.XPATH, f"//*[@id='pn_id_1-table']/tbody/tr[{r}]/td[5]", 5).text
                if sch_filename in old_names:
                    wait_for_clickable_element(By.XPATH, f"//*[@id='pn_id_1-table']/tbody/tr[{r}]/td[9]/p-button/button").click()
                    time.sleep(2)
                    gui.hotkey('enter')
            wait_for_clickable_element(By.XPATH, "//button[@aria-label='Next Page']").click()
driver.quit()

