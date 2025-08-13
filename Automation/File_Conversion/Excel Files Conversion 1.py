import os
import pandas as pd
import xlrd, xlwt
from openpyxl import Workbook, load_workbook
import _functions as cfx
from datetime import datetime
import pyxlsb

def convert_file(source_file, dest_folder, dest_ext):
    try:
        src_ext = os.path.splitext(source_file)[1].lower()
        dest_filenm = os.path.split(source_file)[1].replace(src_ext, dest_ext)
        dest_file = os.path.join(dest_folder, dest_filenm)
        if src_ext == ".xls" and dest_ext == ".xlsx":
            workbook_xls = xlrd.open_workbook(source_file)
            workbook_xlsx = Workbook()

            if "Sheet" in workbook_xlsx.sheetnames:
                del workbook_xlsx["Sheet"]

            for sheet_index in range(workbook_xls.nsheets):
                sheet_xls = workbook_xls.sheet_by_index(sheet_index)
                sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
                for row_idx in range(sheet_xls.nrows):
                    for col_idx in range(sheet_xls.ncols):
                        cell_value = sheet_xls.cell_value(row_idx, col_idx)
                        cell_type = sheet_xls.cell_type(row_idx, col_idx)
                        if cell_type == xlrd.XL_CELL_DATE:
                            date_tuple = xlrd.xldate_as_tuple(cell_value, workbook_xls.datemode)
                            cell_value = datetime(*date_tuple)
                            cell_value = cell_value.strftime("%m/%d/%Y")
                        sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            workbook_xlsx.save(dest_file)

        elif src_ext == ".csv" and dest_ext == ".xlsx":
            try:
                df = pd.read_csv(source_file, encoding="utf-8")
            except UnicodeDecodeError:
                df = pd.read_csv(source_file, encoding="ISO-8859-1")
                
            df.to_excel(dest_file, index=False, engine='openpyxl')

        elif src_ext == ".xlsx" and dest_ext == ".xls":
            date_style = xlwt.XFStyle()
            date_style.num_format_str = 'MM/DD/YYYY'

            workbook_xlsx = load_workbook(source_file)
            workbook_xls = xlwt.Workbook()
            for sheet_name in workbook_xlsx.sheetnames:
                sheet_xlsx = workbook_xlsx[sheet_name]
                sheet_xls = workbook_xls.add_sheet(sheet_name)
                for row_idx, row in enumerate(sheet_xlsx.iter_rows(values_only=True)):
                    for col_idx, cell_value in enumerate(row):
                        if isinstance(cell_value, datetime):
                            sheet_xls.write(row_idx, col_idx, cell_value, date_style)
                        else:
                            sheet_xls.write(row_idx, col_idx, cell_value)
            workbook_xls.save(dest_file)
        
        elif src_ext == ".xlsx" and dest_ext == ".csv":
            excel_data = pd.ExcelFile(source_file)
            for sheet_name in excel_data.sheet_names:
                df = excel_data.parse(sheet_name)
                dest_file = os.path.join(os.path.split(dest_file)[0], f"{sheet_name}.csv")
                df.to_csv(dest_file, index=False)

        elif src_ext == ".xlsx" and dest_ext == ".xlsb":
            workbook_xlsx = load_workbook(source_file)
            workbook_xlsb = xlwt.Workbook()
            for sheet_name in workbook_xlsx.sheetnames:
                sheet_xlsx = workbook_xlsx[sheet_name]
                sheet_xlsb = workbook_xlsb.add_sheet(sheet_name)
                for row_idx, row in enumerate(sheet_xlsx.iter_rows(values_only=True)):
                    for col_idx, cell_value in enumerate(row):
                        if isinstance(cell_value, datetime):
                            sheet_xlsb.write(row_idx, col_idx, cell_value, date_style)
                        else:
                            sheet_xlsb.write(row_idx, col_idx, cell_value)
            workbook_xlsb.save(dest_file)
        elif src_ext == ".xlsb" and dest_ext == ".xlsx":
            workbook_xlsx = Workbook()
            del workbook_xlsx['Sheet']  # remove default sheet

            with pyxlsb.open_workbook(source_file) as wb:
                for sheet_name in wb.sheets:
                    ws_xlsx = workbook_xlsx.create_sheet(title=sheet_name)
                    with wb.get_sheet(sheet_name) as sheet:
                        for row_idx, row in enumerate(sheet.rows()):
                            for col_idx, cell in enumerate(row):
                                value = cell.v
                                if isinstance(value, datetime):
                                    ws_xlsx.cell(row=row_idx + 1, column=col_idx + 1).value = value
                                else:
                                    ws_xlsx.cell(row=row_idx + 1, column=col_idx + 1).value = value

            workbook_xlsx.save(dest_file)
        else:
            raise ValueError(f"Unsupported conversion: {src_ext} to {dest_ext}")

    except Exception as e:
        print(f"Failed to convert {source_file}: {e}")

choice = int(cfx.inputbox(title='Options', prompt='1. Folder\n2. File'))
fl_format = int(cfx.inputbox(title='Conversion', prompt='1. XLS to XLSX \n2. CSV to XLSX\n3. XLSX to XLS\n4. XLSX to CSV\n5. XLSX to XLSB\n6. XLSB to XLSX'))
extensions = {
    1: (".xls", ".xlsx"),
    2: (".csv", ".xlsx"),
    3: (".xlsx", ".xls"),
    4: (".xlsx", ".csv"),
    5: (".xlsx", ".xlsb"),
    6: (".xlsb", ".xlsx")
}
src_ext, dest_ext = extensions.get(fl_format, (None, None))

def batch_convert_files(src_folder, src_ext, dest_ext):
    dest_folder_nm = cfx.inputbox(title='Destination Folder', prompt='Type the name of the Destionation Folder')
    dest_folder = os.path.join(os.path.dirname(src_folder), dest_folder_nm)
    os.makedirs(dest_folder, exist_ok=True)

    for file_name in os.listdir(src_folder):
        if file_name.endswith(src_ext):
            source_file = os.path.join(src_folder, file_name)
            convert_file(source_file, dest_folder, dest_ext)

if choice == 1:
    src_folder = cfx.ifolder()
    batch_convert_files(src_folder, src_ext, dest_ext)

elif choice == 2:
    source_file = cfx.ifile()
    dest_folder_nm = cfx.inputbox(title='Destination Folder', prompt='Type the name of the Destionation Folder')
    dest_folder = os.path.join(os.path.dirname(source_file), dest_folder_nm)
    os.makedirs(dest_folder, exist_ok=True)
    convert_file(source_file, dest_folder, dest_ext)