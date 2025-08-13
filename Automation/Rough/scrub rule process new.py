import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook 
import _functions as cfx
import os
import pyautogui as pag

pag.alert(text='Please Close The Below files if open and Click OK.\n\n\t1. OPDATA.CSV File\n\n\t2. BILLNUM_TAG.XLSX\n\n\t3. Final OPData', title='Close Files', button='OK')

initialdir = r"C:\Users\karthikeyan.s\Documents\Sage Servicewise Report\SCRUBRULE - PYTHON\0Source Files"
opfile_path = cfx.ifile(title='OPDATA.CSV', initialdir=initialdir)
opfile_folder_path = os.path.dirname(opfile_path)
opfile_filename = os.path.basename(opfile_path)
dirname = os.path.split(opfile_folder_path)[0]

opfile_df = pd.read_csv(opfile_path, header=None, low_memory=False)
col_start = int(cfx.inputbox(title='Column Start', prompt='Enter Column Start Number'))
col_list = []
for col in opfile_df.columns:
    xx = opfile_df[col].unique()
    nn = len(xx)
    if nn == 1:
        col_list.append(xx[0])

billnum_col = col_list.index('S.NO')
start = billnum_col - (billnum_col)
icddesc_col = col_list.index('ICDDESC')
end = icddesc_col - billnum_col

req_col_list_nm = ['PAYER']
for i in range(billnum_col, icddesc_col+1):
    req_col_list_nm.append(col_list[i])

req_col_list_idx = [40]
for i in range(start, end+1):
    req_col_list_idx.append(col_start+i)

for col_num, col_name in zip(req_col_list_idx, req_col_list_nm):
    opfile_df.rename({col_num: col_name}, axis=1, inplace=True)

opfile_df = opfile_df[req_col_list_nm]

opfile_path = os.path.join(dirname, "1OP Data", "OP_DATA.xlsx")
    
opfile_df.to_excel(opfile_path, index=False)

op_wb = load_workbook(opfile_path)
op_sh = op_wb.active
cfx.autofit_columns(op_sh)
op_wb.save(opfile_path)

data_path = opfile_path

cl_path = os.path.join(dirname, '2Lookup Files', 'CHECKLIST.xlsx')
nc_path = os.path.join(dirname, '2Lookup Files', 'NON-COVERED.xlsx')

data_df0 = pd.read_excel(data_path, sheet_name=0)
data_df = data_df0.sort_values(by=['S.NO'], ascending=[True])
id_cpt_df = data_df[['BILLNUM','CPTCODE']]
id_cpt_df_grpby = id_cpt_df.groupby('BILLNUM')['CPTCODE'].agg(lambda x: ', '.join(x.astype(str))).reset_index()
id_cpt_df_grpby_dict = id_cpt_df_grpby.set_index('BILLNUM')['CPTCODE'].to_dict()

cl_df = pd.read_excel(cl_path, sheet_name=0)
cl_df_dict = cl_df.set_index('Check')['Tag'].to_dict()
new_wb = Workbook()
new_sh = new_wb.active
cl_tag_dict = {}

nc_df = pd.read_excel(nc_path, sheet_name=0)
nc_df = nc_df[['Code', 'Coverage Status']]
nc_df.rename(columns={'Code': 'ICDCODE'}, inplace=True)

for billnum, cptcode in id_cpt_df_grpby_dict.items():
    # cptcode = str(cptcode)
    cptcode = cptcode.split(', ')
    cptcode = list(set(cptcode))
    for check, tag in cl_df_dict.items():
        check = str(check)
        check = check.split(', ')
        check = list(set(check))
        check_cnt = len(check)
        common_items = list(set(check) & set(cptcode))
        common_items_cnt = len(common_items)
        if common_items and check_cnt == common_items_cnt:
            if not billnum in cl_tag_dict:
                cl_tag_dict[billnum] = []
            cl_tag_dict[billnum].append(tag)
x=2

cl_tag_data = []
for billnum, tag in cl_tag_dict.items():
    tag = " ; ".join(tag)
    cl_tag_data.append([billnum, tag])

cl_tag_df = pd.DataFrame(cl_tag_data, columns=['BILLNUM', 'Tag'])

cl_tag_file_path = os.path.join(dirname, '2Lookup Files', 'BILLNUM_TAG.xlsx')

cl_tag_df.to_excel(cl_tag_file_path, sheet_name='BILLNUM_TAG', index=False)

cl_tag_wb = load_workbook(cl_tag_file_path)
cl_tag_sh = cl_tag_wb['BILLNUM_TAG']
cfx.autofit_columns(cl_tag_sh)
cl_tag_wb.save(cl_tag_file_path)

data_cl_df = pd.merge(data_df, cl_tag_df, on="BILLNUM", how="left")
dups_df = data_cl_df.duplicated(subset="BILLNUM", keep="first")
cl_tag_columns = cl_tag_df.columns.difference(["BILLNUM"])
data_cl_df.loc[dups_df, cl_tag_columns] = np.nan

data_cl_nc_df = pd.merge(data_cl_df, nc_df, on="ICDCODE", how="left")


cptchange_df = data_cl_nc_df[['PATCODE', 'BILLDATE', 'PROVDESC', 'CPTCODE']]
cptchange_df = cptchange_df[(cptchange_df['CPTCODE'].isin(['9', '10', '11'])) | (cptchange_df['CPTCODE'].isin([9, 10, 11]))].reset_index(drop=True)
cptchange_df['unique0'] = cptchange_df['CPTCODE'] + "_" + cptchange_df['PATCODE'] + "_" + cptchange_df['PROVDESC']

cptchange_df['BILLDATE'] = pd.to_datetime(cptchange_df['BILLDATE'], dayfirst=True, errors='coerce')
cptchange_df['UNIQUE'] = cptchange_df['CPTCODE'] + "_" + cptchange_df['PATCODE'] + "_" + cptchange_df['PROVDESC'] + "_" + cptchange_df['BILLDATE'].dt.strftime('%d/%m/%Y')

cptchange_df = cptchange_df.drop_duplicates(subset=['UNIQUE'], keep='first').reset_index(drop=True)

billdate_lst = cptchange_df.groupby(['CPTCODE', 'unique0'])['BILLDATE'].apply(lambda x: ", ".join(x.astype(str).unique())).reset_index()
billdate_lst = billdate_lst.rename(columns={'BILLDATE': 'BILLDATE_Lst'})
cptchange_df = pd.merge(cptchange_df, billdate_lst, on=['CPTCODE', 'unique0'], how='left')
cptchange_df['BILLDATE_Cnt'] = cptchange_df.groupby(['CPTCODE', 'unique0'])['BILLDATE'].transform(lambda x: x.nunique())
cptchange_df['BILLDATE_Min'] = cptchange_df.groupby(['CPTCODE', 'unique0'])['BILLDATE'].transform(lambda x: min(x.dropna().unique()) if len(x.dropna()) > 0 else pd.NaT)

cptchange_df = cptchange_df[cptchange_df['BILLDATE_Cnt'] != 1].reset_index(drop=True)
cptchange_df['BILLDATE_Min'] = pd.to_datetime(cptchange_df['BILLDATE_Min'], dayfirst=True, errors='coerce')
cptchange_df['Days'] = (cptchange_df['BILLDATE'] - cptchange_df['BILLDATE_Min']).dt.days

conitions = [
    (cptchange_df['Days'] >= 1) & (cptchange_df['Days'] <=7),
    (cptchange_df['Days'] >= 8) & (cptchange_df['Days'] <=31),
    (cptchange_df['Days'] == 0)
]

ncpt1 = None
ncpt2 = (cptchange_df['CPTCODE'].astype(int) + .02).astype(str)
ncpt3 = cptchange_df['CPTCODE'].astype(int).astype(str)

ncpt = [ncpt1, ncpt2, ncpt3]

rptvst1 = f"Repeat Visit [Before 7 Days] : " + cptchange_df['PATCODE'] + " - First Visit on " + cptchange_df['BILLDATE_Min'].dt.strftime('%d/%m/%Y')
rptvst2 = f"Repeat Visit [After 7 Days] : " + cptchange_df['PATCODE'] + " - First Visit on " + cptchange_df['BILLDATE_Min'].dt.strftime('%d/%m/%Y')
rptvst3 = None

rptvst = [rptvst1, rptvst2, rptvst3]

cptchange_df['NewCPTCODE1'], cptchange_df['Repeat Visit'] = np.select(conitions, ncpt), np.select(conitions, rptvst)

final_file_path0 = os.path.join(dirname, '2Lookup Files', 'CPTCODE_CHANGE.xlsx')
cptchange_df.to_excel(final_file_path0, sheet_name='CPTCODE_CHANGE', index=False)

cptchange_df = cptchange_df[['UNIQUE', 'NewCPTCODE1', 'Repeat Visit']]

data_cl_nc_df['BILLDATE'] = pd.to_datetime(data_cl_nc_df['BILLDATE'], dayfirst=True, errors='coerce')
data_cl_nc_df['UNIQUE'] = data_cl_nc_df['CPTCODE'] + "_" + data_cl_nc_df['PATCODE'] + "_" + data_cl_nc_df['PROVDESC'] + "_" + data_cl_nc_df['BILLDATE'].dt.strftime('%d/%m/%Y')
final_data_df1 = pd.merge(data_cl_nc_df, cptchange_df, on='UNIQUE', how='left')

def fxnewcpt(row):
    if str(row['CPTCODE']) in ['9', '10', '11']:
        return row['NewCPTCODE1']
    else:
        return row['CPTCODE']

final_data_df1['NewCPTCODE'] = final_data_df1.apply(fxnewcpt, axis=1)

col_order = ['PAYER', 'BILLNUM', 'VISITID', 'BILLDATE', 'PATCODE', 'PATNAME', 'EID', 'POLICYNO', 'SERVICE TYPE', 'BILLCODE', 'BILLDESC', 'CPTCODE', 'NewCPTCODE', 'Repeat Visit', 'CPTDESC', 'PROVDESC', 'VISITDEPT', 'ORDQTY', 'UNITCOST', 'SUBTOTAL', 'DISCAMT', 'NETEXTCOST', 'PATRESP', 'COMPRESP', 'SCOMPRESP', 'APPROVLNO', 'ORDNO', 'LICENSE', 'NETWRKCODE', 'NETWORDESC', 'CLMSTATUS', 'RATLSTCODE', 'RATLSTCODE.1', 'INSU(Y/N)', 'RATLSTCODE.2', 'RATLSTCODE.3', 'DIAGTYP', 'ICDCODE', 'ICDDESC', 'Tag', 'Coverage Status']
final_data_df1 = final_data_df1[col_order]

final_folder = os.path.join(dirname, '3Final Output')
final_filename = opfile_filename.replace('.csv', '.xlsx')
final_file_path = os.path.join(final_folder, final_filename)

final_data_df1.to_excel(final_file_path, sheet_name='OP_DATA', index=False)

final_data_wb = load_workbook(final_file_path)
final_data_sh = final_data_wb['OP_DATA']
cfx.autofit_columns(final_data_sh)
final_data_wb.save(final_file_path)
os.startfile(final_folder)
