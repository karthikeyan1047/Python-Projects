import xlwings as xw
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Side, Border, Font, PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkcalendar import DateEntry
from datetime import datetime, timedelta
import pandas as pd
import os, shutil, stat, time, xlrd, xlwt, csv

# HELPER FUNCTIONS
def detect_delimiter(file_path, encoding="utf-8"):
    with open(file_path, "r", encoding=encoding) as file:
        sample = file.read(1024)
        sniffer = csv.Sniffer()
        try:
            return sniffer.sniff(sample).delimiter
        except Exception:
            return ","

def permission(func, path, exc_info):
    import stat
    os.chmod(path, stat.S_IWRITE)
    func(path)

# FUNCTIONS
def ifile(title='Select File',initialdir=None):
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title=title,
        initialdir=initialdir
    )
    return os.path.normpath(file_path)

def ifolder(title='Select Folder', initialdir=None):
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(
        title=title,
        initialdir=initialdir
    )
    return os.path.normpath(folder_path)

def borders():
    bds=Border(
        left=Side(border_style='thin') , 
        right=Side(border_style='thin') ,
        top=Side(border_style='thin') ,
        bottom=Side(border_style='thin')
    )
    return bds

def clearing(sheet):
    try:
        for row in sheet[sheet.min_row:sheet.max_row]:
            for cell in row:
                cell.value = None
                cell.fill = PatternFill()
                cell.border = Border()
                cell.font = Font()
    except:
        cell = sheet['A1']
        cell.value = None
        cell.fill = PatternFill()
        cell.border = Border()
        cell.font = Font()

def show_info(title, prompt):
    root = tk.Tk()
    root.attributes('-topmost', True)

    window_width = 300
    window_height = 150
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    root.withdraw()
    root.withdraw()
    messagebox.showinfo(title, prompt)
    root.quit()

def yesno(title, prompt):
    root = tk.Tk()
    root.attributes('-topmost', True)

    window_width = 300
    window_height = 150
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")
    
    root.withdraw()
    root.withdraw()
    response = messagebox.askyesno(title, prompt)
    if response:
        return True
    else:
        return False

def close_book(file_path):
    wb = xw.Book(file_path)
    wb.save()
    wb.close()
    xw.App().quit()

def inputbox(title, prompt):
    ttl = title
    pmt = prompt
    root = tk.Tk()
    root.attributes('-topmost', True)

    window_width = 300
    window_height = 150
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    root.withdraw()
    user_input = simpledialog.askstring(ttl, pmt)
    return user_input

def autofit_columns(sheet):
    last_row = sheet.max_row
    last_col = sheet.max_column
    for col in range(1, last_col + 1):
        max_length = 0
        column = openpyxl.utils.get_column_letter(col)
        for row in range(1, last_row + 1):
            cell = sheet[f'{column}{row}']
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    if "_" in str(cell.value):
                        max_length = len(str(cell.value)) + 2
                    else:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 3
        sheet.column_dimensions[column].width = adjusted_width

def get_date(text):
    root = tk.Tk()
    root.title("Date")
    root.attributes('-topmost', True)

    window_width = 125
    window_height = 125
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    label = tk.Label(root, text=text)
    label.pack(pady=10)
    date_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='mm/dd/yyyy')
    date_entry.pack(pady=10)
    selected_date = None
    def on_submit():
        nonlocal selected_date
        selected_date = date_entry.get_date()
        root.quit()
        root.destroy()
    submit_button = tk.Button(root, text="Submit", command=on_submit)
    submit_button.pack(pady=10)
    root.mainloop()
    return selected_date

def get_file_details(folder_path, excel_output):
    if not os.path.exists(folder_path):
        print(f"The folder '{folder_path}' does not exist.")
        return
    file_details = []
    for file in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            created_time = os.path.getctime(file_path)
            modified_time = os.path.getmtime(file_path)
            file_details.append({
                "File Name": file,
                "Created Date": datetime.fromtimestamp(created_time).strftime('%Y-%m-%d %H:%M:%S'),
                "Modified Date": datetime.fromtimestamp(modified_time).strftime('%Y-%m-%d %H:%M:%S')
            })
    df = pd.DataFrame(file_details)
    df.to_excel(excel_output, index=False)

def get_folder_details(main_folder, excel_output):
    if not os.path.exists(main_folder):
        print(f"The folder '{main_folder}' does not exist.")
        return
    folder_details = []
    for folder in os.listdir(main_folder):
        folder_path = os.path.join(main_folder, folder)
        if os.path.isdir(folder_path):
            created_time = os.path.getctime(folder_path)
            modified_time = os.path.getmtime(folder_path)
            folder_details.append({
                "Folder Name": folder,
                "Created Date": datetime.fromtimestamp(created_time).strftime('%Y-%m-%d %H:%M:%S'),
                "Modified Date": datetime.fromtimestamp(modified_time).strftime('%Y-%m-%d %H:%M:%S')
            })
    df = pd.DataFrame(folder_details)
    df.to_excel(excel_output, index=False)

def update_file_details(main_folder, excel_file):
    if not os.path.exists(main_folder):
        print(f"The main folder '{main_folder}' does not exist.")
        return
    df = pd.read_excel(excel_file)
    required_columns = {"File Name", "Created Date", "Modified Date"}
    if not required_columns.issubset(df.columns):
        print(f"The Excel file must contain these columns: {required_columns}")
        return
    for index, row in df.iterrows():
        file_name = row["File Name"]
        created_date = row["Created Date"]
        modified_date = row["Modified Date"]
        created_timestamp = time.mktime(datetime.strptime(created_date, "%Y-%m-%d %H:%M:%S").timetuple())
        modified_timestamp = time.mktime(datetime.strptime(modified_date, "%Y-%m-%d %H:%M:%S").timetuple())
        file_path = os.path.join(main_folder, file_name)
        if os.path.exists(file_path) and os.path.isfile(file_path):
            os.utime(file_path, (created_timestamp, modified_timestamp))
            print(f"Updated file: {file_name}")
        else:
            print(f"File '{file_name}' does not exist in '{main_folder}'.")

def update_folder_details(main_folder, excel_file):
    if not os.path.exists(main_folder):
        print(f"The main folder '{main_folder}' does not exist.")
        return
    df = pd.read_excel(excel_file)
    required_columns = {"Folder Name", "Created Date", "Modified Date"}
    if not required_columns.issubset(df.columns):
        return
    for index, row in df.iterrows():
        folder_name = row["Folder Name"]
        created_date = row["Created Date"]
        modified_date = row["Modified Date"]
        created_timestamp = time.mktime(datetime.strptime(created_date, "%Y-%m-%d %H:%M:%S").timetuple())
        modified_timestamp = time.mktime(datetime.strptime(modified_date, "%Y-%m-%d %H:%M:%S").timetuple())
        folder_path = os.path.join(main_folder, folder_name)
        if os.path.exists(folder_path) and os.path.isdir(folder_path):
            os.utime(folder_path, (created_timestamp, modified_timestamp))
        else:
            print(f"Folder '{folder_name}' does not exist in '{main_folder}'.")


def keep_latest_files(main_folder, keep_count):
    files_to_delete = [
        os.path.join(main_folder, file)
        for file in os.listdir(main_folder)
        if os.path.isfile(os.path.join(main_folder, file))
    ]
    files_to_delete.sort(key=os.path.getctime, reverse=True)
    files_to_delete = files_to_delete[keep_count:]
    for file in files_to_delete:
            os.remove(file)

def keep_latest_folder(main_folder, keep_count):
    def wirte_access(func, path, exc_info):
        os.chmod(path, stat.S_IWRITE)
        func(path)
    flds_to_delete = [
        os.path.join(main_folder, fld) 
        for fld in os.listdir(main_folder) 
        if os.path.isdir(os.path.join(main_folder, fld))
    ]
    flds_to_delete.sort(key=os.path.getctime, reverse=True)
    flds_to_delete = flds_to_delete[keep_count:]
    for f in flds_to_delete:
        shutil.rmtree(f, onerror=wirte_access)

def create_date_folders(start_date, end_date, base_path):
    start_date = datetime.strftime(start_date, "%Y.%m.%d")
    end_date = datetime.strftime(end_date, "%Y.%m.%d")
    start = datetime.strptime(start_date, "%Y.%m.%d")
    end = datetime.strptime(end_date, "%Y.%m.%d")
    current_date = start
    while current_date <= end:
        folder_name = current_date.strftime("%Y.%m.%d")
        folder_path = os.path.join(base_path, folder_name)
        os.makedirs(folder_path, exist_ok=True)
        current_date += timedelta(days=1)

def create_date_files(main_folder, start_date, end_date, extension):
    start_date = datetime.strftime(start_date, "%Y.%m.%d")
    end_date = datetime.strftime(end_date, "%Y.%m.%d")
    start = datetime.strptime(start_date, "%Y.%m.%d")
    end = datetime.strptime(end_date, "%Y.%m.%d")
    current_date = start
    while current_date <= end:
        file_name = current_date.strftime("%Y.%m.%d") + extension
        file_path = main_folder + "/" + file_name
        workbook  = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        workbook.save(file_path)
        current_date += timedelta(days=1)

def create_date_files(main_folder, filenames, extension):
    for fls in filenames:
        file_path = main_folder + "/" + fls + extension
        workbook  = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        workbook.save(file_path)

def create_list_folders(main_folder, foldersnames):
    for flds in foldersnames:
        folder_path = main_folder + "/" + flds
        os.makedirs(folder_path)

def xls_to_xlsx_n(src_folder):
    for file_name in os.listdir(src_folder):
        if file_name.endswith(".xls"):
            source_file = os.path.join(src_folder, file_name)
            dest_folder = os.path.join(src_folder, "XLSX")
            os.makedirs(dest_folder, exist_ok=True)
            destination_file = os.path.join(dest_folder, file_name.replace(".xls", ".xlsx"))
            try:
                workbook_xls = xlrd.open_workbook(source_file)
                workbook_xlsx = Workbook()

                if "Sheet" in workbook_xlsx.sheetnames:
                    del workbook_xlsx["Sheet"]
                
                for sheet_index in range(workbook_xls.nsheets):
                    sheet_xls = workbook_xls.sheet_by_index(sheet_index)
                    sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
                    for row_idx in range(sheet_xls.nrows):
                        for col_idx in range(sheet_xls.ncols):
                            cell_value = sheet_xls.cell_value(row_idx, col_idx)
                            sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                workbook_xlsx.save(destination_file)
            except Exception as e:
                print(f"Failed to convert {source_file}: {e}")
    os.startfile(src_folder)

def xls_to_xlsx(src_file):
    try:
        workbook_xls = xlrd.open_workbook(src_file)
        workbook_xlsx = Workbook()

        if "Sheet" in workbook_xlsx.sheetnames:
            del workbook_xlsx["Sheet"]

        for sheet_index in range(workbook_xls.nsheets):
            sheet_xls = workbook_xls.sheet_by_index(sheet_index)
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
            for row_idx in range(sheet_xls.nrows):
                for col_idx in range(sheet_xls.ncols):
                    cell_value = sheet_xls.cell_value(row_idx, col_idx)
                    sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

        dest_file = os.path.splitext(src_file)[0] + ".xlsx"
        workbook_xlsx.save(dest_file)

    except Exception as e:
        print(f"Failed to convert {src_file}: {e}")

    os.startfile(os.path.split(src_file)[0])

def csv_to_xlsx_n(folder_path, encoding="utf-8"):
    if not os.path.exists(folder_path):
        return
    csv_files = [file for file in os.listdir(folder_path) if file.endswith('.csv')]

    if not csv_files:
        return
    output_folder = os.path.join(folder_path, "XLSX")
    os.makedirs(output_folder, exist_ok=True)

    for csv_file in csv_files:
        csv_path = os.path.join(folder_path, csv_file)
        xlsx_file = os.path.splitext(csv_file)[0] + ".xlsx"
        xlsx_path = os.path.join(output_folder, xlsx_file)

        try:
            try:
                delimiter = detect_delimiter(csv_path, encoding=encoding)
            except UnicodeDecodeError:
                delimiter = detect_delimiter(csv_path, encoding="ISO-8859-1")
            try:
                df = pd.read_csv(csv_path, sep=delimiter, encoding=encoding)
            except UnicodeDecodeError:
                df = pd.read_csv(csv_path, sep=delimiter, encoding="ISO-8859-1")
            
            df.to_excel(xlsx_path, index=False, engine="openpyxl")
        except Exception as e:
            print(f"Error converting '{csv_file}': {e}")
    os.startfile(folder_path)

def csv_to_xlsx(csv_file, encoding="utf-8"):
    try:
        delimiter = detect_delimiter(csv_file, encoding=encoding)
    except UnicodeDecodeError:
        delimiter = detect_delimiter(csv_file, encoding="ISO-8859-1")
    try:
        df = pd.read_csv(csv_file, sep=delimiter, encoding=encoding)
    except UnicodeDecodeError:
        df = pd.read_csv(csv_file, sep=delimiter, encoding="ISO-8859-1")

    xlsx_file = os.path.splitext(csv_file)[0] + ".xlsx"
    df.to_excel(xlsx_file, index=False, engine='openpyxl')
    os.startfile(os.path.split(xlsx_file)[0])

def convert_file_1(source_file, dest_folder, dest_ext):
    try:
        src_ext = os.path.splitext(source_file)[1].lower()
        dest_filenm = os.path.split(source_file)[1].replace(src_ext, dest_ext)
        dest_file = os.path.join(dest_folder, dest_filenm)
        if src_ext == ".xls" and dest_ext == ".xlsx":
            workbook_xls = xlrd.open_workbook(source_file)
            workbook_xlsx = Workbook()

            if "Sheet" in workbook_xlsx.sheetnames:
                del workbook_xlsx["Sheet"]

            for sheet_index in range(workbook_xls.nsheets):
                sheet_xls = workbook_xls.sheet_by_index(sheet_index)
                sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
                for row_idx in range(sheet_xls.nrows):
                    for col_idx in range(sheet_xls.ncols):
                        cell_value = sheet_xls.cell_value(row_idx, col_idx)
                        cell_type = sheet_xls.cell_type(row_idx, col_idx)
                        if cell_type == xlrd.XL_CELL_DATE:
                            date_tuple = xlrd.xldate_as_tuple(cell_value, workbook_xls.datemode)
                            cell_value = datetime(*date_tuple)
                            cell_value = cell_value.strftime("%m/%d/%Y")
                        sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            workbook_xlsx.save(dest_file)

        elif src_ext == ".csv" and dest_ext == ".xlsx":
            try:
                df = pd.read_csv(source_file, encoding="utf-8")
            except UnicodeDecodeError:
                df = pd.read_csv(source_file, encoding="ISO-8859-1")
                
            df.to_excel(dest_file, index=False, engine='openpyxl')

        elif src_ext == ".xlsx" and dest_ext == ".xls":
            date_style = xlwt.XFStyle()
            date_style.num_format_str = 'MM/DD/YYYY'

            workbook_xlsx = load_workbook(source_file)
            workbook_xls = xlwt.Workbook()
            for sheet_name in workbook_xlsx.sheetnames:
                sheet_xlsx = workbook_xlsx[sheet_name]
                sheet_xls = workbook_xls.add_sheet(sheet_name)
                for row_idx, row in enumerate(sheet_xlsx.iter_rows(values_only=True)):
                    for col_idx, cell_value in enumerate(row):
                        if isinstance(cell_value, datetime):
                            sheet_xls.write(row_idx, col_idx, cell_value, date_style)
                        else:
                            sheet_xls.write(row_idx, col_idx, cell_value)
            workbook_xls.save(dest_file)
        
        elif src_ext == ".xlsx" and dest_ext == ".csv":
            excel_data = pd.ExcelFile(source_file)
            for sheet_name in excel_data.sheet_names:
                df = excel_data.parse(sheet_name)
                dest_file = os.path.join(os.path.split(dest_file)[0], f"{sheet_name}.csv")
                df.to_csv(dest_file, index=False)

        else:
            raise ValueError(f"Unsupported conversion: {src_ext} to {dest_ext}")

    except Exception as e:
        print(f"Failed to convert {source_file}: {e}")

def convert_files(src_folder, src_ext, dest_ext):
    dest_folder = os.path.join(src_folder, dest_ext[1:].upper())
    os.makedirs(dest_folder, exist_ok=True)

    for file_name in os.listdir(src_folder):
        if file_name.endswith(src_ext):
            source_file = os.path.join(src_folder, file_name)
            convert_file_1(source_file, dest_folder, dest_ext)
