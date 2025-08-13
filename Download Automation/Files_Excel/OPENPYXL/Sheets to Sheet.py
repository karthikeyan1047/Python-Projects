import _functions as cfx
import openpyxl, os, shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

file_path = cfx.ifile()
workbook = load_workbook(file_path)
combined_book = Workbook()
combined_sheet = combined_book.active
combined_sheet.title = "Combined"

is_first_sheet = True
current_row = 1
date_style = NamedStyle(name='date_style', number_format="MM/DD/YYYY")
if "date_style" not in combined_book.named_styles:
    combined_book.add_named_style(date_style)

for sheet in workbook.sheetnames:
    current_sheet = workbook[sheet]
    for row_num, row in enumerate(current_sheet.iter_rows(min_row=1, max_row=current_sheet.max_row, min_col=1, max_col=current_sheet.max_column), start=1):
        if row_num == 1 and not is_first_sheet:
            continue 
        data = [cell.value for cell in row]
        for col_num, value in enumerate(data, start=1):
            if isinstance(value, datetime):
                combined_sheet.cell(row=current_row, column=col_num, value=value).style = date_style
            else:
                combined_sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1
    is_first_sheet = False

dest_directory = os.path.dirname(file_path)
dest_folder = os.path.join(dest_directory, "Sheets_Sheet")

if os.path.exists(dest_folder):
    shutil.rmtree(dest_folder)
os.makedirs(dest_folder)

dest_file = os.path.join(dest_folder, "Sheets_Sheet.xlsx")
combined_book.save(dest_file)
os.startfile(dest_folder)