import openpyxl, os, shutil
import _functions as cfx
from datetime import datetime
from openpyxl.styles import NamedStyle

file_path = cfx.ifile()
folder_path = os.path.join(os.path.dirname(file_path), "Sheets_Files")
if os.path.exists(folder_path):
    shutil.rmtree(folder_path)
os.makedirs(folder_path)
date_style = NamedStyle(name='date_style', number_format="MM/DD/YYYY")
workbook = openpyxl.load_workbook(file_path)

for sheet_name in workbook.sheetnames:
    current_row = 1
    new_workbook = openpyxl.Workbook()
    sheet = workbook[sheet_name]
    new_sheet = new_workbook.active
    new_sheet.title = sheet_name
    for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column), start=1):
        data = [cell.value for cell in row]
        for col_num, value in enumerate(data, start=1):
            if isinstance(value, datetime):
                new_sheet.cell(row=current_row, column=col_num, value=value).style = date_style
            else:
                new_sheet.cell(row=current_row, column=col_num, value=value)
        current_row += 1    
    
    new_filename = os.path.join(folder_path, f"{sheet_name}.xlsx")
    new_workbook.save(new_filename)

os.startfile(folder_path)

