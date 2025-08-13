import os, shutil, sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
import _functions as cfx
from datetime import datetime


choice = int(cfx.inputbox(title='Method', prompt='Choose the suitable method\n   1. Single File_Single Sheet\n   2. Single File_Multiple Sheets\n   3. Multiple Files_Multiple Sheets'))

###########################################################
### SINGLE FILE WITH SINGLE SHEET - SPLIT DATA BY CATEGORY
###########################################################

if choice == 1:
    file_path = cfx.ifile()
    filter_column = int(cfx.inputbox(title='Column Number', prompt="Enter the column number contains category to split"))

    folder_path = os.path.dirname(file_path)
    odirectory = os.path.join(folder_path, "Sheet to Files")
    if os.path.exists(odirectory):
        shutil.rmtree(odirectory)
    os.makedirs(odirectory)

    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[0]
    date_style = NamedStyle(name='date_style', number_format="MM/DD/YYYY")
    categories = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        category = row[filter_column-1]
        if category not in categories:
            categories[category] = []
        categories[category].append(row)

        headers = [cell.value for cell in sheet[1]]

        for category, rows in categories.items():
            dest_file = Workbook()
            new_sheet = dest_file.active
            category = " ".join(category.split(" ")[:2])
            new_sheet.title = str(category)

            for col_num, header in enumerate(headers, start=1):
                new_sheet.cell(row=1, column=col_num, value=header)
            
            for row_num, row_data in enumerate(rows, start=2):
                for col_num, cell_value in enumerate(row_data, start=1):
                    if isinstance(cell_value, datetime):
                        new_sheet.cell(row=row_num, column=col_num, value=cell_value).style = date_style
                    else:
                        new_sheet.cell(row=row_num, column=col_num, value=cell_value)

            dest_file_path = os.path.join(odirectory, f"{str(category)}.xlsx")
            dest_file.save(dest_file_path)
        
    os.startfile(os.path.dirname(odirectory))

##########################################################################
### SINGLE FILE WITH MULTIPLE SHEETS - SPLIT AND COMBINE DATA BY CATEGORY
##########################################################################

elif choice == 2:
    file_path = cfx.ifile()
    filter_column = int(cfx.inputbox(title='Column Number', prompt="Enter the column number contains category to split"))

    folder_path = os.path.dirname(file_path)
    odirectory = os.path.join(folder_path, "Sheet to Files")
    if os.path.exists(odirectory):
        shutil.rmtree(odirectory)
    os.makedirs(odirectory)

    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[0]
    date_style = NamedStyle(name='date_style', number_format="MM/DD/YYYY")
    categories = {}

    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            category = row[filter_column-1]
            if category not in categories:
                categories[category] = []
            categories[category].append(row)

    for category, rows in categories.items():
        dest_file = Workbook()
        new_sheet = dest_file.active
        category = " ".join(category.split(" ")[:2])
        new_sheet.title = str(category)

        for col_num, header in enumerate(headers, start=1):
            new_sheet.cell(row=1, column=col_num, value=header)
        
        for row_num, row_data in enumerate(rows, start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                if isinstance(cell_value, datetime):
                    new_sheet.cell(row=row_num, column=col_num, value=cell_value).style = date_style
                else:
                    new_sheet.cell(row=row_num, column=col_num, value=cell_value)

        dest_file_path = os.path.join(odirectory, f"{str(category)}.xlsx")
        dest_file.save(dest_file_path)
        
    os.startfile(os.path.dirname(odirectory))


#############################################################################
### MULTIPLE FILES WITH MULTIPLE SHEETS - SPLIT AND COMBINE DATA BY CATEGORY
#############################################################################

elif choice == 3:
    folder_path = cfx.ifolder()
    filter_column = int(cfx.inputbox(title='Column Number', prompt="Enter the column number contains category to split"))

    odirectory = os.path.join(folder_path, "Sheet to Files")
    if os.path.exists(odirectory):
        shutil.rmtree(odirectory)
    os.makedirs(odirectory)

    date_style = NamedStyle(name='date_style', number_format="MM/DD/YYYY")
    categories = {}

    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        workbook = load_workbook(file_path)
        for sheet in workbook.worksheets:
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                category = row[filter_column-1]
                if category not in categories:
                    categories[category] = []
                categories[category].append(row)

    for category, rows in categories.items():
        dest_file = Workbook()
        new_sheet = dest_file.active
        category = " ".join(category.split(" ")[:2])
        new_sheet.title = str(category)

        for col_num, header in enumerate(headers, start=1):
            new_sheet.cell(row=1, column=col_num, value=header)
        
        for row_num, row_data in enumerate(rows, start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                if isinstance(cell_value, datetime):
                    new_sheet.cell(row=row_num, column=col_num, value=cell_value).style = date_style
                else:
                    new_sheet.cell(row=row_num, column=col_num, value=cell_value)

        dest_file_path = os.path.join(odirectory, f"{str(category)}.xlsx")
        dest_file.save(dest_file_path)
        
    os.startfile(os.path.dirname(odirectory))

else:
    sys.exit(0)