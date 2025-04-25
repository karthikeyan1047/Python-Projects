from selenium import webdriver
import _functions as cfx
import os, sys
import xlwings as xw
from openpyxl.styles import Font, Border, Side, PatternFill, NamedStyle
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import time

fill = PatternFill(start_color="ffcc00", end_color="ffcc00", fill_type="solid")
initialdir = r"C:\Users\karthikeyan.s\Desktop\Processmed"
workbook_path = cfx.ifile(initialdir=initialdir)

book_ext = os.path.splitext(os.path.basename(workbook_path))[1]

if book_ext == '.xlsm':
    workbook = load_workbook(workbook_path, keep_vba=True)
elif book_ext == '.xlsx':
    workbook = load_workbook(workbook_path)
else:
    sys.exit(0)

soa_sh = workbook['SOA']
cfx.close_book(workbook_path)

options = Options()
options.add_argument("--start-maximized")
options.add_argument("--disable-popup-blocking")

options.add_experimental_option('detach', True)
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(20)

url = "https://rcm.processmed.ae:5070/"
iusername = "HealthBay"
ipassword = "*9D0TNC:4%a_ghug"
chtemplate = "HealthBay_SOA"
xinstances = ["Motor City", "Womens Care", "Day Surgery", "Verve"]
rows = [5, 20, 35, 50]
yrfrm = 2023
yrto = datetime.now().year
enstart = "Encounter.Start"

driver.get(url)
driver.maximize_window()

# Login
driver.find_element(By.ID, "username").clear()
driver.find_element(By.ID, "username").send_keys(iusername)
driver.find_element(By.ID, "password").clear()
driver.find_element(By.ID, "password").send_keys(ipassword)
driver.find_element(By.XPATH, "//a[contains(@class, 'btn btn-primary') and text()=' Sign in']").click()

cfx.clearing(soa_sh)

Select(driver.find_element(By.ID, "selexporttemp")).select_by_visible_text(chtemplate)
Select(driver.find_element(By.ID, "selparametersdate")).select_by_visible_text(enstart)

for oo in range(0,4):
    Select(driver.find_element(By.ID, "selclientname")).select_by_index(oo)
    for yr in range(yrfrm, yrto + 1):
        moninp = datetime.now().month if yr == yrto else 12

        for i in range(1, moninp + 1):
            date_from = datetime(yr, i, 1).date().strftime("%m/%d/%Y")
            if i < 12:
                date_to = (datetime(yr, i + 1, 1) - timedelta(days=1)).strftime("%m/%d/%Y")
            else:
                date_to = (datetime(yr + 1, 1, 1) - timedelta(days=1)).strftime("%m/%d/%Y")

            driver.find_element(By.ID, "filterSubTransDateFrm1").clear()
            driver.find_element(By.ID, "filterSubTransDateFrm1").send_keys(date_from)

            driver.find_element(By.ID, "filterSubTransDateTo1").clear()
            driver.find_element(By.ID, "filterSubTransDateTo1").send_keys(date_to)

            # (-) and (+) Click
            driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr[2]/td[2]/a[2]").click()
            driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr[2]/td[2]/a[1]").click()

            # (Schedule) Click
            driver.find_element(By.XPATH, "//*[@id='event-container']/form/fieldset[2]/a[1]").click()
            WebDriverWait(driver, 10).until(EC.alert_is_present())
            Alert(driver).accept()

            # (Report) Click
            driver.find_element(By.XPATH, "//*[@id='event-container']/form/fieldset[2]/a[2]").click()
            driver.refresh()

            # Sort by created date in decending order
            driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/thead/tr/th[5]").click()
            driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/thead/tr/th[5]").click()

            itext = driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/tbody/tr[1]/td[5]").text[-11:]
            colno = yr - 2022
            cell_row = rows[oo]

            cell_value = f"[ {datetime(yr, i, 1).date().strftime('%b')} - {yr} ] - [ {itext} ]"

            if soa_sh.cell(row=cell_row, column=colno).value is None:
                dcell = soa_sh.cell(row=cell_row, column=colno)
                dcell.value = cell_value
                dcell.border = cfx.borders()
            else:
                dcell = soa_sh.cell(row=cell_row + i-1, column=colno)
                dcell.value = cell_value
                dcell.border = cfx.borders()

            driver.back()

    dwl_date = datetime.today().strftime("%d/%m/%Y")
    dwl_date = f"Dowmnload Date : {dwl_date}"
    soa_sh.cell(row=1, column=1).value=dwl_date
    soa_sh.cell(row=1, column=1).font = Font(bold=True)

    soa_sh.cell(row=rows[oo] - 1, column=1).value = xinstances[oo]
    soa_sh.cell(row=rows[oo] - 1, column=1).border = cfx.borders()
    soa_sh.cell(row=rows[oo] - 1, column=1).fill = fill
    
    cfx.autofit_columns(soa_sh)

workbook.save(workbook_path)
driver.find_element(By.XPATH, "//*[@id='logout']/span/a").click()
driver.quit()

wb = xw.Book(workbook_path)
wb.sheets["SOA"].activate()