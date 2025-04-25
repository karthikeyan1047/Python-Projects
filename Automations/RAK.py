from selenium import webdriver
import _functions as cfx
import xlwings as xw
import os, sys, datetime, time
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

initialdir = r"C:\Users\karthikeyan.s\Desktop\Processmed"
workbook_path = cfx.ifile(initialdir=initialdir)
cfx.close_book(workbook_path)

# workbook_path = r"C:\Users\karthikeyan.s\Desktop\Processmed\WebAutomation_ProcessMed.xlsx"

rak_wb = load_workbook(workbook_path)
rak_sh = rak_wb["RAK"]

cfx.clearing(rak_sh)
fill = PatternFill(start_color="ffcc00", end_color="ffcc00", fill_type="solid")

options = Options()
options.add_argument("--start-maximized")
options.add_argument("--disable-popup-blocking")

driver = webdriver.Chrome(options=options)
driver.implicitly_wait(20)

def wait_for_element(by, value, timeout=20):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))

def login(url, username, password):
    driver.get(url)
    wait_for_element(By.ID, "username").clear()
    driver.find_element(By.ID, "username").send_keys(username)
    driver.find_element(By.ID, "password").clear()
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.XPATH, "//a[contains(@class, 'btn btn-primary') and text()=' Sign in']").click()
    return username

def select_template(template_name, date_param):
    wait_for_element(By.ID, "selexporttemp").click()
    driver.find_element(By.XPATH, f"//option[text()='{template_name}']").click()
    driver.find_element(By.ID, "selparametersdate").click()
    driver.find_element(By.XPATH, f"//option[text()='{date_param}']").click()

def scheduling(date_from, date_to):
    driver.find_element(By.ID, "filterSubTransDateFrm1").clear()
    driver.find_element(By.ID, "filterSubTransDateFrm1").send_keys(date_from)
    driver.find_element(By.ID, "filterSubTransDateTo1").clear()
    driver.find_element(By.ID, "filterSubTransDateTo1").send_keys(date_to)

    # (-) then (+) Click
    driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr[2]/td[2]/a[2]").click()
    driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr[2]/td[2]/a[1]").click()

    # (Schedule) Click & Accept Alert
    driver.find_element(By.XPATH, "//*[@id='event-container']/form/fieldset[2]/a[1]").click()
    WebDriverWait(driver, 10).until(EC.alert_is_present())
    Alert(driver).accept()

    # (Reports) Click
    driver.find_element(By.XPATH, "//*[@id='event-container']/form/fieldset[2]/a[2]").click()
    driver.refresh()

    # Sort by created date in descending order
    driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/thead/tr/th[5]").click()
    driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/thead/tr/th[5]").click()

    # Get scheduled_date
    return driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/tbody/tr[1]/td[5]").text[-11:]

def process_hospital_pharmacy(yr_start, sheet_start_row, template, iheader):
    select_template(template, "Encounter.End")
    yr_end = datetime.date.today().year
    x = 1
    for yr in range(yr_start, yr_end + 1):
        wait_for_element(By.ID, "selclientname").click()
        driver.find_element(By.XPATH, f"//option[text()='{rak_hosp}']").click()

        months = datetime.date.today().month if yr == datetime.date.today().year else 12

        for i in range(1, months + 1):
            date_from = datetime.date(yr, i, 1).strftime("%m/%d/%Y")
            date_to = (datetime.date(yr, i + 2, 1) - datetime.timedelta(days=1)).strftime("%m/%d/%Y") if i+2 < 12 else datetime.date(yr, 12, 31).strftime("%m/%d/%Y")

            scheduled_date = scheduling(date_from, date_to)

            # Write to Excel
            if template == "Resub_Remittance":
                colno = (yr_end - yr_start) + 3 + yr - (yr_start - 1)
                if x == 1:
                    hcol = colno
                    x += 1
            else:
                colno = yr - (yr_start - 1)
                hcol = 1

            header_cell = rak_sh.cell(row=sheet_start_row, column=hcol)
            header_cell.value = iheader
            header_cell.border = cfx.borders()
            header_cell.fill = fill
            data_cell = rak_sh.cell(row=sheet_start_row + i , column=colno)
            data_cell.value = f"Hosp [ {datetime.date(yr, i, 1).strftime('%b')} - {yr} ] - [ {scheduled_date} ]"
            data_cell.border = cfx.borders()

            # Go back and reset
            driver.back()
            driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr[2]/td[2]/a[2]").click()

        driver.find_element(By.ID, "selclientname").click()
        driver.find_element(By.XPATH, f"//option[text()='{rak_ph}']").click()
        date_from = datetime.date(yr, 1, 1).strftime("%m/%d/%Y")
        if yr == datetime.date.today().year:
            moninp = datetime.date.today().month
            date_to = (datetime.date(yr, moninp+1, 1) - datetime.timedelta(days=1)).strftime("%m/%d/%Y")
        else:
            date_to = datetime.date(yr, 12, 31).strftime("%m/%d/%Y")

        scheduled_date = scheduling(date_from, date_to)

        count = 0
        for r in range(1, rak_sh.max_row+1):
            if (rak_sh.cell(row=r, column=colno)).value != None:
                count = r
        rak_sh.cell(row=count+1, column=colno).value = f"Ph [ {yr} ] - [ {scheduled_date} ]"
        rak_sh.cell(row=count+1, column=colno).border = cfx.borders()

        # Go back and reset
        driver.back()
        driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr[2]/td[2]/a[2]").click()
    
    dwl_date = datetime.date.today().strftime("%d/%m/%Y")
    dwl_date = f"Dowmnload Date : {dwl_date}"
    rak_sh.cell(row=1, column=1, value=dwl_date)

# Process DHPO
user_nm = login("https://rcm.processmed.ae:5070/", "rak1", "SyEfU150>t~aYpv*")
rak_hosp = "Ras Al Khaimah Hospital"
rak_ph = "RAK Pharmacy"
process_hospital_pharmacy(2022, 3, "Unsettled-SMCL", "DHPO - HIS")
process_hospital_pharmacy(2023, 22, "Remittance_", "DHPO - FPA")
process_hospital_pharmacy(2023, 22, "Resub_Remittance", "DHPO - Linewise")

# Process RPO
user_nm = login("https://rcm.processmed.ae:5070/", "rakriyati", "vg5OP?1:/6:A99GK")
rak_hosp = "RAK HOSPITAL"
rak_ph = "RAK HOSPITAL PHARMACY"
process_hospital_pharmacy(2022, 42, "Unsettled-SMCL", "RPO - HIS")
process_hospital_pharmacy(2023, 62, "Remittance_", "RPO - FPA")
process_hospital_pharmacy(2023, 62, "Resub_Remittance", "RPO - Linewise")

cfx.autofit_columns(rak_sh)
rak_wb.save(workbook_path)

driver.quit()
wb = xw.Book(workbook_path)
wb.sheets["RAK"].activate()