from selenium import webdriver
import _functions as cfx
import os, sys, time
import xlwings as xw
from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from datetime import timedelta

def setup_driver():
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-popup-blocking")
    options.add_experimental_option('detach', True)
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(20)
    return driver

def login(driver, url, username, password):
    driver.get(url)
    driver.maximize_window()
    driver.find_element(By.ID, "username").clear()
    driver.find_element(By.ID, "username").send_keys(username)
    driver.find_element(By.ID, "password").clear()
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.XPATH, "//*[@id='login-form']/footer/a").click()

def setup_filters(driver, template, status, date_from, date_to, last_transaction_date):
    Select(driver.find_element(By.ID, "selexporttemp")).select_by_visible_text(template)
    # Select(driver.find_element(By.ID, "selparameters")).select_by_visible_text("status")
    # Select(driver.find_element(By.ID, "selcondition")).select_by_visible_text("Equals")
    # driver.find_element(By.ID, "filtervalue").clear()
    # driver.find_element(By.ID, "filtervalue").send_keys(status)
    # driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr/td[4]/a[2]").click()
    # driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr/td[4]/a[1]").click()

    Select(driver.find_element(By.ID, "selparametersdate")).select_by_visible_text(last_transaction_date)
    driver.find_element(By.ID, "filterSubTransDateFrm1").clear()
    driver.find_element(By.ID, "filterSubTransDateFrm1").send_keys(date_from)
    driver.find_element(By.ID, "filterSubTransDateTo1").clear()
    driver.find_element(By.ID, "filterSubTransDateTo1").send_keys(date_to)
    driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr[2]/td[2]/a[2]").click()
    driver.find_element(By.XPATH, "//*[@id='inbox-table']/tbody/tr[2]/td[2]/a[1]").click()

def scheduling(driver):
    driver.find_element(By.XPATH, "//*[@id='event-container']/form/fieldset[2]/a[1]").click()
    WebDriverWait(driver, 10).until(EC.alert_is_present())
    Alert(driver).accept()
    driver.find_element(By.XPATH, "//*[@id='event-container']/form/fieldset[2]/a[2]").click()
    driver.refresh()
    driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/thead/tr/th[5]").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/thead/tr/th[5]").click()
    itext = driver.find_element(By.XPATH, "//*[@id='datatable_fixed_column']/tbody/tr[1]/td[5]").text[-11:]
    return itext

def header_styles(cell, font, fill):
    cell.font = font
    cell.fill = fill
    cell.border = cfx.borders()

def data_styles(cell):
    cell.border = cfx.borders()

def process_clients(driver, rec_sh, clients, flnames, start_row):
    for i, client in enumerate(clients):
        Select(driver.find_element(By.ID, "selclientname")).select_by_visible_text(client)
        itext = scheduling(driver)
        rec_sh.cell(row=start_row + i, column=1).value = flnames[i]
        rec_sh.cell(row=start_row + i, column=1).border = cfx.borders()
        rec_sh.cell(row=start_row + i, column=2).value = client
        rec_sh.cell(row=start_row + i, column=2).border = cfx.borders()
        rec_sh.cell(row=start_row + i, column=3).value = itext
        rec_sh.cell(row=start_row + i, column=3).border = cfx.borders()
        driver.back()

fnt = Font(bold=True)
orange_fill = PatternFill(start_color="ffcc00", end_color="ffcc00", fill_type="solid")
blue_fill = PatternFill(start_color="ffcc00", end_color="ffcc00", fill_type="solid")

if not cfx.yesno("Process", "Want to Proceed or Not?"):
    sys.exit()

initialdir = r"C:\Users\karthikeyan.s\Desktop\Processmed"
workbook_path = cfx.ifile(initialdir=initialdir)
book_ext = os.path.splitext(os.path.basename(workbook_path))[1]

if book_ext not in ['.xlsm', '.xlsx']:
    sys.exit()

workbook = load_workbook(workbook_path, keep_vba=(book_ext == '.xlsm'))
rec_sh = workbook['Reconciliation']
cfx.close_book(workbook_path)
cfx.clearing(rec_sh)

url = "https://rcm.processmed.ae:5070/"
datefrm = cfx.get_date("Data From : ")
dateto = cfx.get_date("Date To : ")
datediff = (dateto - datefrm).days // 3
mdate1 = datefrm + timedelta(days=datediff)
mdate2 = datefrm + timedelta(days=2*datediff)
dates = [datefrm, mdate1, mdate2, dateto]
datefrm, dateto, mdate1, mdate2 = datefrm.strftime("%m/%d/%Y"), dateto.strftime("%m/%d/%Y"), mdate1.strftime("%m/%d/%Y"), mdate2.strftime("%m/%d/%Y")
print(datefrm)
print(dateto)

driver = setup_driver()
login(driver, url, "rak2", "7k9{U:s7*}Zv-BmD")
setup_filters(driver, "Re-submission_Audit1", "Remitted", datefrm, dateto, "lastTransactionDate")

clients = ["RAK HOSPITAL", "RAK HOSPITAL PHARMACY", "RMC AL HAMRA", "Ras Al Khaimah Medical Center Galilah", "RAS AL KHAIMAH MEDICAL CENTRE AL JAZEERA", 
            "Modern Medical Laboratory", "Shiyas and Ifthikar Medical Center Riyati", "RMC AL JAZEERA PHARMACY", "Medstar", "Rak Medical Center - Al Hamra", 
            "Rak Medical Center - Jazeera", "Rak Medical Center - Ghalila", "MMLN", "Star Metropolis Clinic Laboratory", "SHIYAS AND IFTHIKAR MEDICAL CENTER"]
flnames = ["BI_RPO_IPOP", "BI_RPO_PH", "BI_RPO_L7026", "BI_RPO_L5223", "BI_RPO_L5136", "BI_RPO_L5741", "BI_RPO_L5887", 
            "BI_RPO_L206", "XML_BI_med", "XML_BI_HOC", "XML_BI_JOC", "XML_BI_GOC", "XML_BI_MML", "XML_BI_SMCL", "XML_BI_443"]

process_clients(driver, rec_sh, clients, flnames, 2)

hcell = rec_sh['A1']
hcell.value = "BI_RPO"
header_styles(cell=hcell, font=fnt, fill=orange_fill)

cfx.autofit_columns(rec_sh)
driver.back()

# driver = setup_driver()
login(driver, url, "rak1", "SyEfU150>t~aYpv*")
clients1 = ["Ras Al Khaimah Hospital", "RAK Pharmacy"]
flnames1 = ["BI_DHPO_IPOP", "BI_DHPO_PH"]
cr = 20
x = 1
for i, client in enumerate(clients1):
    for j in range(len(dates) - 1):
        Select(driver.find_element(By.ID, "selclientname")).select_by_visible_text(client)
        dtfrm = dates[j].strftime("%m/%d/%Y") if j == 0 else (dates[j] + timedelta(days=1)).strftime("%m/%d/%Y")
        dtto = dates[j + 1].strftime("%m/%d/%Y")
        setup_filters(driver, "Re-submission_Audit1", "Remitted", dtfrm, dtto, "lastTransactionDate")
        itext = scheduling(driver)
        hcell = rec_sh.cell(row=cr, column=1)
        hcell.value = "BI_DHPO"
        header_styles(cell=hcell, font=fnt, fill=orange_fill)

        lists = [f"{flnames1[i]} ", client, itext]
        for o, val in enumerate(lists):
            dcell = rec_sh.cell(row=cr + x, column=o+1)
            dcell.value = lists[o]
            data_styles(cell=dcell)
        driver.back()
        x += 1
        
cfx.autofit_columns(rec_sh)
driver.quit()

drs_dhpo1 = ["XML_DRS_SMCL_DHA", "XML_DRS_JOC", "XML_DRS_HOC", "XML_DRS_med", "XML_DRS_simc", "XML_DRS_SMCL_MOH", "XML_DRS_GOC","DRS_DHPO_IPOP", "DRS_DHPO_PH"]
drs_dhpo2 = ["Smcl", "rmcjazeera", "alhamra-rmc", "Medstar2015", "SIMCNEW", "smclsharjah", "rakmc-ghalilla", "rakhospital", "rakhpharmacy"] 
drs_rpo1 = ["DRS_RPO_L153", "DRS_RPO_L5223", "DRS_RPO_L5741", "DRS_RPO_L7026", "DRS_RPO_L5887", "DRS_RPO_L5136", "DRS_RPO_L206 ", "DRS_RPO_IPOP ", "DRS_RPO_PH "]
drs_rpo2 = ["L153", "L5223", "L5741", "L7026", "L5887", "L5136", "L206", "L5067", "L1145"]

rec_sh['F1'].value = "Reconciliation"
rec_sh['F1'].fill = blue_fill
rec_sh['F1'].border = cfx.borders()

def fill_reconciliation_data(start_row, data1, data2, label):
    rec_sh[f"F{start_row}"].value = label
    rec_sh[f"F{start_row}"].font = fnt
    rec_sh[f"F{start_row}"].fill = orange_fill
    for i, (d1, d2) in enumerate(zip(data1, data2), start=1):
        rec_sh.cell(row=start_row + i, column=6).value = d1
        rec_sh.cell(row=start_row + i, column=7).value = d2
        rec_sh.cell(row=start_row + i, column=6).border = cfx.borders()
        rec_sh.cell(row=start_row + i, column=7).border = cfx.borders()

fill_reconciliation_data(4, drs_dhpo1, drs_dhpo2, "DRS_DHPO")
fill_reconciliation_data(len(drs_dhpo1) + 7, drs_rpo1, drs_rpo2, "DRS_RPO")

rec_sh.cell(row=1, column=11).value = "Default Count"
rec_sh.cell(row=1, column=12).value = len(clients) + 6 + len(drs_dhpo1) + len(drs_rpo1) + 3

cfx.autofit_columns(rec_sh)
workbook.save(workbook_path)

wb = xw.Book(workbook_path)
wb.sheets["Reconciliation"].activate()